"""Kenya Food Systems and Land Use Action Plan (2024-2030): policy-layer extract.

The Action Plan is a fixed, published government PDF. This module:

  1. locates the PDF in data/raw/action_plan/ ,
  2. records provenance (SHA-256, size, page count, extraction time),
  3. extracts the full text to data/processed/action_plan/action_plan_fulltext.txt
     (pdfplumber preferred, pypdf fallback) so nothing in the source is lost,
  4. verifies the curated tables below against that text (anchor check),
  5. writes the human workbook "Kenya action plan structured.xlsx" to
     data/processed/action_plan/ , and
  6. writes tidy per-table CSVs to data/external/action_plan/ , which the
     generic ingester folds into the policy layer on the next build.

The budget table, appendices and figure callouts in this PDF use merged and
multi-line cells that no general table parser reads reliably, so the structured
tables here are a curated transcription of the document, checked by formula
(the 7-year plan reconciles to 52,700 KES millions) and cross-checked against
the extracted text. The PDF itself supplies the checksum, page count and full
text for the provenance record.

Author: Aboubacar HEMA
"""
from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path

RQ = "\u2019"  # typographic apostrophe


def _t(s: str) -> str:
    """Enforce the project typographic standard on text fields: no straight
    apostrophes, no en/em dashes."""
    return s.replace("'", RQ).replace("\u2013", "-").replace("\u2014", "-")


WB_NAME = "Kenya action plan structured.xlsx"

# --------------------------------------------------------------------------- #
# Curated source data (verbatim figures from the Action Plan, October 2023)
# --------------------------------------------------------------------------- #

# Table 3, full plan. Each block: (critical_transition, [(item, priority, cost_kes_m), ...])
BUDGET_7YR = [
    ("Healthy diets", [
        ("i", "Develop and implement policies to foster food diversification and de-risking credit and insurance.", 1000.00),
        ("ii", "Promote production, consumption, preservation and trade of diversified and nutrition-adequate diets (especially plant-rich diets).", 2400.00),
        ("iii", "Focus on the implementation of a coherent, conducive legal, regulatory and institutional framework for plant seeds, animal breeds and fingerlings production, multiplication, distribution and marketing of the outputs.", 1200.00),
        ("iv", "Scale-up programs to promote and facilitate fortification or bio-fortification of widely consumed staple food.", 1700.00),
        ("v", "Promote the utilisation and consumption of forgotten/orphan/indigenous foods.", 1200.00),
        ("vi", "Review and implement micronutrient, healthy diet guidelines and strategies.", 1500.00),
    ]),
    ("Productive and regenerative agriculture", [
        ("i", "Review and implement a SMART fertiliser Subsidy Programme to make it more focused, cost-effective and efficient.", 2150.00),
        ("ii", "Promote sustainable practices in agricultural production, good soil health, biodiversity and better land use options.", 2400.00),
        ("iii", "Climate-proofing food production and land use systems through climate change agriculture practices and insurance.", 2200.00),
        ("iv", "Promoting the use of diversified extension systems (VBA, MOA).", 2000.00),
    ]),
    ("Protecting and restoring nature", [
        ("i", "Promote implementation of a coherent, conducive legal, regulatory and institutional framework for natural resources and strengthening land tenure and governance.", 1300.00),
        ("ii", "Map land use potential for alternative land uses and relook at the land tenure system (long-term).", 2400.00),
        ("iii", "Boost nature-positive production.", 1800.00),
        ("iv", "Promote an integrated approach to natural resource management.", 1200.00),
        ("v", "Undertake capacity building for communities to enable them to manage land use effectively.", 2200.00),
    ]),
    ("Food loss and waste", [
        ("i", "Invest more in nutrition education and create SMART regulations for food safety without overburdening businesses, at both national and county levels.", 1600.00),
        ("ii", "Review and harmonise the current laws and regulations to address emerging challenges of food safety and provide for functional responsibilities along the food value chain at both national and county levels.", 1400.00),
        ("iii", "Develop and maintain modern physical market infrastructure (assembly or wholesale, retail and rural markets) with cold storage, cleaning, sorting and grading sheds, value addition and processing facilities for efficient marketing of agricultural produce.", 2400.00),
        ("iv", "Implement actions to reduce waste and losses in production and post-harvest operations.", 1200.00),
        ("v", "Undertake sensitisation and awareness creation activities on FLW.", 1200.00),
        ("vi", "Implement actions to reduce waste and losses in production and post-harvest operations and undertake sensitisation and awareness creation activities on FLW.", 1000.00),
        ("vii", "Facilitate compliance to enhance compliance with set agricultural, food produce and products standards.", 800.00),
        ("viii", "Enhance the capacity of surveillance institutions at national and county levels to enforce compliance with standards.", 1400.00),
        ("ix", "Harmonise and strengthen inter-agency efforts in food safety and quality control and monitoring to minimise overlaps and inter-agency conflicts.", 700.00),
        ("x", "Develop an FLW protocol and capacity-build stakeholders on reducing food loss and waste.", 400.00),
    ]),
    ("Gender and social inclusion", [
        ("i", "Invest in capacity building for youth (particularly in agribusiness), producers (extension system) and women along the value chains, and set up clear contract enforcement mechanisms that protect small-scale farmers.", 1375.00),
        ("ii", "Promote youth and women participation and entrepreneurship in the food and land use system.", 1375.00),
        ("iii", "Promote enhanced youth access to productive resources and inputs.", 1500.00),
        ("iv", "Promote women/youth-led food and land use research, innovation and technology adoption.", 850.00),
        ("v", "Promote access to markets and value addition for women/youth-led enterprises.", 850.00),
        ("vi", "Strengthen youth and women leadership, capacity development and coordination in the food system.", 750.00),
    ]),
    ("Monitoring and evaluation", [
        ("", "Monitoring and evaluation of the action plan.", 3250.00),
    ]),
    ("Coordination of strategy implementation", [
        ("", "Coordination meetings for the action plan implementation.", 4000.00),
    ]),
]

# 2024 priority areas. The document prints Total = 1,600 although the listed
# sub-totals sum to 3,750; the figures are kept as printed and the discrepancy
# is surfaced (workbook data note + DOC_TOTAL_2024 here).
BUDGET_2024 = [
    ("Healthy diets", [
        ("", "Scale-up programs to promote and facilitate fortification or bio-fortification of widely consumed staple foods.", 400.00),
        ("", "Review and implement micronutrient, healthy diet guidelines and strategies.", 350.00),
    ]),
    ("Productive and regenerative agriculture", [
        ("", "Review and implement a SMART fertiliser Subsidy Programme to make it more focused, cost-effective and efficient.", 450.00),
        ("", "Promoting the use of diversified extension systems (VBA, MOA).", 400.00),
    ]),
    ("Protecting and restoring nature", [
        ("", "Promote an integrated approach to natural resource management.", 250.00),
        ("", "Undertake capacity building for communities to enable them to manage land use effectively.", 250.00),
    ]),
    ("Food loss and waste", [
        ("", "Undertake sensitisation and awareness creation activities on FLW.", 600.00),
        ("", "Implement actions to reduce waste and losses in production and post-harvest operations and undertake sensitisation and awareness creation activities on FLW.", 200.00),
        ("", "Develop an FLW protocol and capacity-build stakeholders on reducing food loss and waste.", 50.00),
    ]),
    ("Gender and social inclusion", [
        ("", "Promote access to markets and value addition for women/youth-led enterprises.", 200.00),
        ("", "Strengthen youth and women leadership, capacity development and coordination in the food system.", 150.00),
        ("", "Invest in capacity building for youth (particularly in agribusiness), producers (extension system) and women along the value chains, and set up clear contract enforcement mechanisms that protect small-scale farmers.", 150.00),
    ]),
    ("Monitoring and evaluation", [
        ("", "Monitoring and evaluation of the action plan.", 100.00),
    ]),
    ("Coordination of strategy implementation", [
        ("", "Coordination meetings for the action plan implementation.", 200.00),
    ]),
]
DOC_TOTAL_2024 = 1600.00

# Child nutrition (under 5) cited in the plan. county = "National" for national.
NUTRITION = [
    ("National", "Stunting", 18, "KDHS 2022", "Too short for age (chronic undernutrition)"),
    ("National", "Wasting", 5, "KDHS 2022", "Too thin for height (acute undernutrition)"),
    ("National", "Underweight", 10, "KDHS 2022", "Too thin for age"),
    ("National", "Overweight", 3, "KDHS 2022", "Sign of overnutrition"),
    ("National", "Stunting", 26, "KDHS 2014", "Trend reference"),
    ("National", "Underweight", 11, "KDHS 2014", "Trend reference"),
    ("National", "Wasting", 4, "KDHS 2014", "Trend reference"),
    ("National", "Stunting", 35, "2008/2009", "Trend reference"),
    ("National", "Stunting", 40, "1993", "Trend reference"),
    ("Kilifi", "Stunting", 37, "KDHS 2022", "Highest county stunting cited"),
    ("West Pokot", "Stunting", 34, "KDHS 2022", "High county stunting cited"),
    ("Samburu", "Stunting", 31, "KDHS 2022", "High county stunting cited"),
    ("Kisumu", "Stunting", 9, "KDHS 2022", "Lowest county stunting cited"),
]

AG_GROWTH_YEARS = ["2018", "2019", "2020", "2021", "2022", "2023"]
AG_GROWTH = [6.1, 2.8, 5.2, -0.3, -1.9, 3.0]  # %; 2023 projected
AG_COMMODITY_YEARS = ["2017", "2018", "2019", "2020", "2021"]
AG_COMMODITIES = [
    ("Maize, total marketed (000 tonnes)", [239.2, 441.5, 316.7, 261.3, 228.4]),
    ("Milk sold centrally (million litres)", [591.4, 652.3, 685.9, 684.4, 801.9]),
    ("Sugar (000 tonnes)", [4715.6, 5262.2, 4606.1, 6810.9, 7783.3]),
    ("Wheat (000 tonnes)", [156.9, 330.3, 348.8, 280.8, 241.9]),
]

TRANSITIONS = [
    ("i", "Healthy diets", "Access to adequate, nutrition-diversified, safe and affordable food for all Kenyans."),
    ("ii", "Productive and regenerative agriculture", "Contributes to sustainable and sufficient production and food supply systems; and to integrated natural resource management and biodiversity restoration."),
    ("iii", "Protecting and restoring nature", "Promote implementation of a conducive policy and legal framework for natural resources and strengthening land tenure and governance."),
    ("iv", "Food loss and waste", "Contributes to sustained reduction of food loss and waste."),
    ("v", "Youth and social inclusion", "Contribute to sustainable food security, increased incomes and inclusive economic growth through gender, youth and social inclusion in the food system and land use."),
]
FOLU_PILLARS = ["Nutritious food", "Nature-based solutions", "Wider choice and supply", "Opportunity for all"]
FOLU_TEN = [
    "Healthy diets", "Productive and regenerative agriculture", "Protecting and restoring nature",
    "A healthy and productive ocean", "Diversifying protein supply", "Reducing food loss and waste",
    "Local loops and linkages", "Harnessing the digital revolution", "Stronger rural livelihoods",
    "Gender and demography",
]

POLICIES = [
    ("Food Policy (Sessional Paper No. 4)", "1981", "1"),
    ("National Food Policy (Sessional Paper 1)", "1994", "1"),
    ("Poverty Reduction Strategy Paper", "2001", "1"),
    ("Economic Recovery Strategy for Wealth and Employment Creation", "2003-2007", "1"),
    ("Kenya Rural Development Strategy (KRDS)", "2002-2017", "1"),
    ("Strategy for Revitalising Agriculture (SRA)", "2004-2014", "1"),
    ("Kenya Vision 2030 and Medium Term Plans (1st, 2nd, 3rd)", "2008-", "1"),
    ("Millennium Development Goals (domestication and ratification)", "", "1"),
    ("Sustainable Development Goals (domestication and ratification)", "", "1"),
    ("Comprehensive African Agriculture Development Programme (CAADP) of NEPAD", "2002", "1"),
    ("National Nutrition Policy", "2008", "1"),
    ("Food Security and Nutrition Policy", "2011", "1"),
    ("Agriculture Sector Development Strategy (ASDS)", "2010-2020", "1"),
    ("National Agribusiness Strategy", "2012", "1"),
    ("National Food Safety Policy", "2013", "1"),
    ("Food Security Bill", "2017", "1"),
    ("Kenya Nutrition Action Plan", "2018-2022", "1"),
    ("Agricultural Sector Transformation and Growth Strategy (ASTGS)", "2019-2029", "1"),
    ("ASTGS National Agriculture Investment Plan (NAIP)", "", "1"),
    ("National Food and Nutrition Security Policy Implementation Framework", "2017-2022", "1"),
    ("National Agricultural Policy", "", "1"),
    ("National Livestock Policy", "", "1"),
    ("National Root and Tuber Crops Development Strategy", "2019-2022", "1"),
    ("National Potato Strategy", "2016-2020", "1"),
    ("National Agricultural Soil Management Policy (NASMP)", "", "2"),
    ("Land Use Policy", "", "2"),
    ("National Irrigation Policy", "", "2"),
    ("National Forest Policy", "", "2"),
    ("Sessional Paper No. 1 on KWS Policy", "", "2"),
    ("Range Management and Pastoralism Strategy", "2021-2031", "2"),
    ("IGAD Regional Post Harvest Loss Management Strategy (IGAD-PHLMS)", "", "2"),
    ("National Climate Change Action Plan", "", "2"),
    ("National Climate Change Response Strategy (NCCRS)", "", "2"),
    ("Kenya National Adaptation Plan", "2015-2030", "2"),
    ("Nationally Determined Contribution", "2015", "2"),
    ("Youth Climate Action Strategy for Kenya (YCASK)", "2021-2030", "2"),
    ("National Agricultural Research System Policy (NARS)", "2021", "2"),
    ("National Agricultural Sector Extension Policy (NASEP)", "", "2"),
]

LEGISLATION = [
    "Enact legislation to make soil liming mandatory",
    "Enact legislation to cap the cost of leasing land to attract private and foreign investors",
    "Enact legislation to halt further subdivision of arable land",
    "Enact Warehouse Receipt System Bill 2016",
    "Enforce Fisheries Management and Development Act",
    "Enforce road legislation to eliminate multiple levies across counties",
    "Food Security Bill, 2014",
    "Legislation to stimulate water harvesting across the country",
    "Legislation on irrigated land for each constituency",
    "Restoration of commodity levies to beef up commodity fund",
    "Legislation on caged fish farming",
    "Enforcement of agriculture regulations (Crops: tea, sugar, potatoes)",
    "Regulations on commodity levies (Sugar Regulations)",
]

_DP, _MIN, _RES, _PRIV, _CTY, _CSO, _OFF = (
    "Development partner", "Ministry/Government", "Research/Academic institution",
    "Private sector", "County government", "CSO/Association", "Official/Individual")
STAKEHOLDERS = [
    ("Alliance for a Green Revolution in Africa (AGRA)", _DP),
    ("Global Alliance for Improved Nutrition (GAIN)", _DP),
    ("World Resources Institute (WRI)", _DP),
    ("Welthungerhilfe (WHH)", _DP),
    ("Food and Agriculture Organization (FAO)", _DP),
    ("Ministry of Agriculture and Livestock Development (MOALD)", _MIN),
    ("Ministry of Lands and Physical Planning (MoLPP)", _MIN),
    ("Ministry of Water and Irrigation (MoW&I)", _MIN),
    ("Ministry of Health (MoH)", _MIN),
    ("University of Nairobi (UoN)", _RES),
    ("TEGEMEO Institute", _RES),
    ("Kenya National Farmers Federation (KENAFF)", _CSO),
    ("Kenya Forestry Research Institute (KEFRI)", _RES),
    ("African Conservation Tillage (ACT)", _CSO),
    ("Agriculture Sector Network (ASNET)", _PRIV),
    ("Seed Trade Association of Kenya (STAK)", _PRIV),
    ("International Tree Foundation (ITF)", _CSO),
    ("Kenya Private Sector Alliance (KEPSA)", _PRIV),
    ("Kenya Land Alliance (KLA)", _CSO),
    ("Kenya Livestock Producers Association (KLPA)", _CSO),
    ("Kenyan Youth Biodiversity Network (KYBN)", _CSO),
    ("Micro-Enterprises Support Programme Trust (MESPT)", _DP),
    ("National Potato Council of Kenya (NPCK)", _CSO),
    ("Association of Women in Agriculture Kenya (AWAK)", _CSO),
    ("One Tree Planted (OTP)", _CSO),
    ("Participatory Ecological Land Use Management (PELUM)", _CSO),
    ("Society of Crop Agribusiness Advisory (SoCAA)", _CSO),
    ("The Organic Movement", _CSO),
    ("County Executive Member, Agriculture, Kisumu", _CTY),
    ("HelloTractor", _PRIV),
    ("Kenya Agricultural and Livestock Research Organization (KALRO)", _RES),
    ("County Executive Member, Agriculture, Kisii", _CTY),
    ("County Executive Member, Agriculture, Homa Bay", _CTY),
    ("Transparency International (TI)", _CSO),
    ("Agro-dealers from Kisumu", _PRIV),
    ("Farmers from Kisumu and Homa Bay counties", _OFF),
    ("County Executive Member, Agriculture, Turkana", _CTY),
    ("County Executive Member, Agriculture, Elgeyo Marakwet", _CTY),
    ("County Executive Member, Agriculture, Uasin Gishu", _CTY),
    ("County Executive Member, Agriculture, Makueni", _CTY),
    ("County Executive Member, Agriculture, Machakos", _CTY),
    ("County Executive Member, Agriculture, Kitui", _CTY),
    ("County Executive Member, Agriculture, Kiambu (rep)", _CTY),
    ("County Executive Member, Agriculture, Kirinyaga (rep)", _CTY),
    ("County Executive Member, Agriculture, Laikipia", _CTY),
    ("County Executive Member, Roads, Laikipia", _CTY),
    ("County Executive Member, Agriculture, Nyandarua (rep)", _CTY),
    ("County Executive Member, Agriculture, Mombasa", _CTY),
    ("KEMRI Nutritionist, Kilifi", _RES),
    ("Nutritionist, Kilifi County", _CTY),
    ("Fisheries Officer, Mombasa County", _CTY),
    ("Hand in Hand East Africa", _CSO),
    ("CEO, Decent Conservation", _CSO),
    ("Representative, NEMA Mombasa", _MIN),
    ("CEO, Jumuia ya Kaunti za Pwani", _CTY),
    ("Tea Executive, Sanica", _PRIV),
    ("Managing Director, Mardin Group", _PRIV),
    ("Chairman, CPF", _OFF),
    ("Director, Green Leaf", _PRIV),
    ("Chairman, KAFOSA", _OFF),
    ("Sub-County Nutrition and Dietetics Coordinator, Mombasa", _CTY),
    ("Crops Officer, Mombasa", _CTY),
    ("Communication Officer, Decent Conservation", _CSO),
    ("Younger Farmer, Mombasa", _OFF),
    ("Fisheries Officer, Mombasa", _CTY),
    ("Young African Workers", _CSO),
    ("Lecturer, Technical University Mombasa", _RES),
    ("Community CBOs", _CSO),
    ("AWAK IT Manager", _OFF),
]

# Distinctive strings that must appear in the extracted PDF text. Used to
# confirm the curated tables describe the document actually on disk.
VERIFY_ANCHORS = [
    "Food Systems and Land Use", "52,700", "52.70 billion", "FOLU", "AGRA",
    "Kilifi", "West Pokot", "Samburu", "Kisumu", "stunted", "KDHS",
    "SMART fertiliser", "Big 4", "NASMP",
]


# --------------------------------------------------------------------------- #
# PDF handling: locate, checksum, full text
# --------------------------------------------------------------------------- #
def find_pdf(base: Path) -> Path | None:
    """Return the Action Plan PDF in data/raw/action_plan/, or None."""
    d = base / "data" / "raw" / "action_plan"
    if not d.exists():
        return None
    pdfs = sorted(d.glob("*.pdf")) + sorted(d.glob("*.PDF"))
    return pdfs[0] if pdfs else None


def _sha256(path: Path) -> str:
    import hashlib
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def extract_fulltext(pdf: Path) -> tuple[str | None, int]:
    """Extract the PDF text. Returns (text, n_pages). (None, 0) if no reader
    library is available. pdfplumber is preferred; pypdf is the fallback."""
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf)) as doc:
            pages = [(p.extract_text() or "") for p in doc.pages]
        return "\n".join(pages), len(pages)
    except ImportError:
        pass
    except Exception as exc:  # noqa: BLE001
        print(f"[action_plan] pdfplumber failed ({exc}); trying pypdf")
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf))
        pages = [(pg.extract_text() or "") for pg in reader.pages]
        return "\n".join(pages), len(pages)
    except Exception:  # noqa: BLE001
        return None, 0


def verify(text: str | None) -> list[str]:
    """Return the list of anchor strings NOT found in the extracted text.
    An empty list means the curated tables match the document on disk.
    Whitespace is normalised on both sides so anchors are not missed when the
    PDF wraps a phrase across lines (e.g. "West\\nPokot")."""
    if not text:
        return list(VERIFY_ANCHORS)
    import re
    low = re.sub(r"\s+", " ", text.lower())
    return [a for a in VERIFY_ANCHORS
            if re.sub(r"\s+", " ", a.lower()) not in low]


# --------------------------------------------------------------------------- #
# Workbook (human deliverable)
# --------------------------------------------------------------------------- #
def build_workbook(out_path: Path) -> Path:
    """Write the multi-sheet "Kenya action plan structured.xlsx". Sub-totals
    and totals are Excel formulas; author metadata is set to Aboubacar HEMA."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    NAVY, BLUE, GREEN, ROW, SUB = "1F3864", "2E75B6", "548235", "EEF3F9", "DDE9F3"
    HEAD_F = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    BODY_F = Font(name="Arial", size=10)
    BOLD_F = Font(name="Arial", size=10, bold=True)
    TITLE_F = Font(name="Arial", size=14, bold=True, color=NAVY)
    NOTE_F = Font(name="Arial", size=9, italic=True, color="595959")
    thin = Side(style="thin", color="C9C9C9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP = Alignment(wrap_text=True, vertical="top")
    CTR = Alignment(horizontal="center", vertical="center")
    RIGHT = Alignment(horizontal="right")

    wb = Workbook()
    wb.properties.creator = "Aboubacar HEMA"
    wb.properties.lastModifiedBy = "Aboubacar HEMA"
    wb.properties.title = "Kenya Food Systems and Land Use Action Plan (2024-2030) - structured extract"

    def header(ws, row, ncols, fill=NAVY):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HEAD_F
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = WRAP
            cell.border = BORDER

    def put(ws, r, c, v, font=BODY_F, fill=None, align=WRAP, num=None, border=True):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font
        cell.alignment = align
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        if num:
            cell.number_format = num
        if border:
            cell.border = BORDER
        return cell

    # README
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    put(ws, 1, 1, "Kenya Food Systems and Land Use Action Plan (2024-2030)", font=TITLE_F, border=False)
    ws.merge_cells("A1:B1")
    meta = [
        ("Document", "Kenya Food Systems and Land Use Action Plan (2024-2030)"),
        ("Publisher", "Government of Kenya, Ministry of Agriculture and Livestock Development; AGRA; FOLU Kenya"),
        ("Date", "October 2023"),
        ("Source role", "Policy layer, structured covariates for the Kenya soil-food-nutrition-policy database"),
        ("Extracted by", "Aboubacar HEMA"),
        ("Note", _t("Figures are transcribed verbatim from the source document. Sub-totals and totals are recomputed by formula; where the document's stated total disagrees with the arithmetic (2024 budget table), the discrepancy is flagged in that sheet.")),
        ("Currency", "KES millions unless stated otherwise"),
    ]
    r = 3
    for k, v in meta:
        put(ws, r, 1, k, font=BOLD_F, border=False)
        put(ws, r, 2, v, border=False)
        r += 1

    def budget_sheet(name, title, blocks, doc_total=None):
        ws = wb.create_sheet(name)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 78
        ws.column_dimensions["D"].width = 16
        put(ws, 1, 1, title, font=TITLE_F, border=False)
        ws.merge_cells("A1:D1")
        hdr = 3
        for i, h in enumerate(["Critical Transition", "Item", "Strategic Priority", "Cost (KES millions)"], 1):
            ws.cell(row=hdr, column=i, value=h)
        header(ws, hdr, 4)
        r = hdr + 1
        sub_rows = []
        for trans, items in blocks:
            first = r
            for code, prio, cost in items:
                put(ws, r, 1, _t(trans) if r == first else None)
                put(ws, r, 2, code, align=CTR)
                put(ws, r, 3, _t(prio))
                put(ws, r, 4, cost, align=CTR, num="#,##0.00")
                r += 1
            if first < r:
                ws.merge_cells(start_row=first, start_column=1, end_row=r - 1, end_column=1)
                ws.cell(row=first, column=1).alignment = WRAP
            put(ws, r, 1, None, fill=SUB)
            put(ws, r, 2, None, fill=SUB)
            put(ws, r, 3, "Sub-Total", font=BOLD_F, fill=SUB, align=RIGHT)
            put(ws, r, 4, f"=SUM(D{first}:D{r-1})", font=BOLD_F, fill=SUB, align=CTR, num="#,##0.00")
            sub_rows.append(r)
            r += 1
        put(ws, r, 1, None, fill=GREEN)
        put(ws, r, 2, None, fill=GREEN)
        put(ws, r, 3, "Total", font=HEAD_F, fill=GREEN, align=RIGHT)
        put(ws, r, 4, "=" + "+".join(f"D{x}" for x in sub_rows), font=HEAD_F, fill=GREEN, align=CTR, num="#,##0.00")
        if doc_total is not None:
            r += 2
            put(ws, r, 1, "Data note", font=BOLD_F, border=False)
            put(ws, r, 3, _t(f"Document states a total of {doc_total:,.2f} KES millions for this table; "
                             "the sum of the listed sub-totals differs. Figures are kept as printed in the "
                             "source; the formula total above reflects the arithmetic of the line items."),
                font=NOTE_F, border=False)
        ws.freeze_panes = f"A{hdr+1}"

    budget_sheet("Budget_7yr", "Table 3. Resource requirements, full plan (2024-2030)", BUDGET_7YR)
    budget_sheet("Budget_2024", "Resource requirements, 2024 priority areas", BUDGET_2024, doc_total=DOC_TOTAL_2024)

    # County_Nutrition
    ws = wb.create_sheet("County_Nutrition")
    for col, w in zip("ABCDE", [18, 16, 12, 16, 42]):
        ws.column_dimensions[col].width = w
    put(ws, 1, 1, "Child nutrition figures cited in the Action Plan (children under 5)", font=TITLE_F, border=False)
    ws.merge_cells("A1:E1")
    hdr = 3
    for i, h in enumerate(["Geography", "Indicator", "Value (%)", "Source", "Note"], 1):
        ws.cell(row=hdr, column=i, value=h)
    header(ws, hdr, 5)
    r = hdr + 1
    for i, (g, ind, val, src, note) in enumerate(NUTRITION):
        fill = ROW if i % 2 else None
        put(ws, r, 1, g, fill=fill)
        put(ws, r, 2, ind, fill=fill)
        put(ws, r, 3, val, fill=fill, align=CTR, num="0")
        put(ws, r, 4, src, fill=fill)
        put(ws, r, 5, _t(note), fill=fill)
        r += 1
    ws.freeze_panes = f"A{hdr+1}"

    # Ag_Performance
    ws = wb.create_sheet("Ag_Performance")
    put(ws, 1, 1, "Table 1. Agricultural sector growth (%), 2018-2023", font=TITLE_F, border=False)
    ws.merge_cells("A1:G1")
    ws.cell(row=3, column=1, value="Indicator")
    for i, y in enumerate(AG_GROWTH_YEARS, 2):
        ws.cell(row=3, column=i, value=y)
    header(ws, 3, 7)
    put(ws, 4, 1, "Agricultural sector growth (%)")
    for i, v in enumerate(AG_GROWTH, 2):
        put(ws, 4, i, v, align=CTR, num="0.0")
    put(ws, 5, 1, "Source: KNBS, National Treasury. 2023 projected. Document text cites a 0.1% contraction in 2021 against the -0.3% shown here.", font=NOTE_F, border=False)
    ws.merge_cells("A5:G5")
    put(ws, 7, 1, "Table 2. Production of selected food commodities, 2017-2021", font=TITLE_F, border=False)
    ws.merge_cells("A7:G7")
    ws.cell(row=9, column=1, value="Commodity (unit)")
    for i, y in enumerate(AG_COMMODITY_YEARS, 2):
        ws.cell(row=9, column=i, value=y)
    header(ws, 9, 6)
    r = 10
    for i, (nm, vals) in enumerate(AG_COMMODITIES):
        fill = ROW if i % 2 else None
        put(ws, r, 1, nm, fill=fill)
        for j, v in enumerate(vals, 2):
            put(ws, r, j, v, fill=fill, align=CTR, num="#,##0.0")
        r += 1
    put(ws, r + 1, 1, "Source: KNBS, 2021.", font=NOTE_F, border=False)
    ws.column_dimensions["A"].width = 36
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 11

    # Critical_Transitions
    ws = wb.create_sheet("Critical_Transitions")
    for col, w in zip("ABC", [8, 40, 80]):
        ws.column_dimensions[col].width = w
    put(ws, 1, 1, "Kenya critical transitions and goals", font=TITLE_F, border=False)
    ws.merge_cells("A1:C1")
    hdr = 3
    for i, h in enumerate(["No.", "Critical Transition", "Goal / objective"], 1):
        ws.cell(row=hdr, column=i, value=h)
    header(ws, hdr, 3)
    r = hdr + 1
    for i, (n, nm, goal) in enumerate(TRANSITIONS):
        fill = ROW if i % 2 else None
        put(ws, r, 1, n, fill=fill, align=CTR)
        put(ws, r, 2, nm, fill=fill)
        put(ws, r, 3, _t(goal), fill=fill)
        r += 1
    r += 1
    put(ws, r, 1, "FOLU four pillars", font=BOLD_F, border=False)
    r += 1
    for p in FOLU_PILLARS:
        put(ws, r, 2, p, border=False)
        r += 1
    r += 1
    put(ws, r, 1, "FOLU ten global transitions (Figure 1)", font=BOLD_F, border=False)
    r += 1
    for p in FOLU_TEN:
        put(ws, r, 2, p, border=False)
        r += 1

    # Policy_Inventory
    ws = wb.create_sheet("Policy_Inventory")
    for col, w in zip("ABCD", [6, 70, 18, 16]):
        ws.column_dimensions[col].width = w
    put(ws, 1, 1, "Policy and strategy inventory (Appendices 1 and 2)", font=TITLE_F, border=False)
    ws.merge_cells("A1:D1")
    hdr = 3
    for i, h in enumerate(["No.", "Policy / strategic document", "Year or period", "Appendix"], 1):
        ws.cell(row=hdr, column=i, value=h)
    header(ws, hdr, 4)
    r = hdr + 1
    for i, (nm, yr, ap) in enumerate(POLICIES, 1):
        fill = ROW if i % 2 == 0 else None
        put(ws, r, 1, i, fill=fill, align=CTR)
        put(ws, r, 2, _t(nm), fill=fill)
        put(ws, r, 3, yr, fill=fill, align=CTR)
        put(ws, r, 4, ap, fill=fill, align=CTR)
        r += 1
    ws.freeze_panes = f"A{hdr+1}"

    # Legislative_Support
    ws = wb.create_sheet("Legislative_Support")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 90
    put(ws, 1, 1, "Legislative support to Big 4 on food and nutrition security (Appendix 3)", font=TITLE_F, border=False)
    ws.merge_cells("A1:B1")
    hdr = 3
    ws.cell(row=hdr, column=1, value="No.")
    ws.cell(row=hdr, column=2, value="Legislative action")
    header(ws, hdr, 2)
    r = hdr + 1
    for i, a in enumerate(LEGISLATION, 1):
        fill = ROW if i % 2 == 0 else None
        put(ws, r, 1, i, fill=fill, align=CTR)
        put(ws, r, 2, _t(a), fill=fill)
        r += 1
    ws.freeze_panes = f"A{hdr+1}"

    # Stakeholders
    ws = wb.create_sheet("Stakeholders")
    for col, w in zip("ABC", [6, 62, 28]):
        ws.column_dimensions[col].width = w
    put(ws, 1, 1, "Stakeholders consulted (Appendix 4)", font=TITLE_F, border=False)
    ws.merge_cells("A1:C1")
    hdr = 3
    for i, h in enumerate(["No.", "Stakeholder", "Category"], 1):
        ws.cell(row=hdr, column=i, value=h)
    header(ws, hdr, 3)
    r = hdr + 1
    for i, (nm, cat) in enumerate(STAKEHOLDERS, 1):
        fill = ROW if i % 2 == 0 else None
        put(ws, r, 1, i, fill=fill, align=CTR)
        put(ws, r, 2, _t(nm), fill=fill)
        put(ws, r, 3, cat, fill=fill)
        r += 1
    ws.freeze_panes = f"A{hdr+1}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# Tidy CSVs (ingestion inputs for the policy layer)
# --------------------------------------------------------------------------- #
def _write_csv(path: Path, header: list[str], rows: list[list]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    return path


def write_csvs(ext_dir: Path) -> list[Path]:
    """Write one tidy CSV per table into data/external/action_plan/. These are
    what the generic ingester turns into policy-layer tables on the next build.
    The county-keyed nutrition CSV also joins to the crosswalk."""
    out: list[Path] = []

    def budget_rows(blocks, period):
        rows = []
        for trans, items in blocks:
            for code, prio, cost in items:
                rows.append([period, _t(trans), code, _t(prio), cost])
        return rows

    out.append(_write_csv(
        ext_dir / "action_plan_budget_7yr.csv",
        ["period", "critical_transition", "item", "strategic_priority", "cost_kes_millions"],
        budget_rows(BUDGET_7YR, "2024-2030")))
    out.append(_write_csv(
        ext_dir / "action_plan_budget_2024.csv",
        ["period", "critical_transition", "item", "strategic_priority", "cost_kes_millions"],
        budget_rows(BUDGET_2024, "2024")))
    out.append(_write_csv(
        ext_dir / "action_plan_county_nutrition.csv",
        ["county", "indicator", "value_pct", "source", "note"],
        [[g, ind, val, src, _t(note)] for g, ind, val, src, note in NUTRITION if g != "National"]))
    out.append(_write_csv(
        ext_dir / "action_plan_national_nutrition.csv",
        ["indicator", "value_pct", "source", "note"],
        [[ind, val, src, _t(note)] for g, ind, val, src, note in NUTRITION if g == "National"]))
    out.append(_write_csv(
        ext_dir / "action_plan_ag_growth.csv",
        ["year", "ag_sector_growth_pct"],
        [[y, v] for y, v in zip(AG_GROWTH_YEARS, AG_GROWTH)]))
    comm_rows = []
    for nm, vals in AG_COMMODITIES:
        for y, v in zip(AG_COMMODITY_YEARS, vals):
            comm_rows.append([nm, y, v])
    out.append(_write_csv(
        ext_dir / "action_plan_ag_commodities.csv",
        ["commodity_unit", "year", "value"], comm_rows))
    out.append(_write_csv(
        ext_dir / "action_plan_critical_transitions.csv",
        ["item", "critical_transition", "goal_objective"],
        [[n, nm, _t(goal)] for n, nm, goal in TRANSITIONS]))
    out.append(_write_csv(
        ext_dir / "action_plan_policy_inventory.csv",
        ["no", "policy_document", "year_or_period", "appendix"],
        [[i, _t(nm), yr, ap] for i, (nm, yr, ap) in enumerate(POLICIES, 1)]))
    out.append(_write_csv(
        ext_dir / "action_plan_legislative_support.csv",
        ["no", "legislative_action"],
        [[i, _t(a)] for i, a in enumerate(LEGISLATION, 1)]))
    out.append(_write_csv(
        ext_dir / "action_plan_stakeholders.csv",
        ["no", "stakeholder", "category"],
        [[i, _t(nm), cat] for i, (nm, cat) in enumerate(STAKEHOLDERS, 1)]))
    return out


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def run(base: Path, *, prov=None) -> dict:
    """Extract the Action Plan and write all local outputs.

    Workbook  -> data/processed/action_plan/<WB_NAME>
    Full text -> data/processed/action_plan/action_plan_fulltext.txt
    Manifest  -> data/processed/action_plan/action_plan_provenance.json
    CSVs      -> data/external/action_plan/*.csv  (ingested into the policy layer)

    Returns a summary dict. Always writes the structured outputs; the PDF is
    used for the checksum, page count, full text and the anchor verification.
    """
    base = Path(base)
    proc = base / "data" / "processed" / "action_plan"
    ext = base / "data" / "external" / "action_plan"
    proc.mkdir(parents=True, exist_ok=True)
    ext.mkdir(parents=True, exist_ok=True)

    pdf = find_pdf(base)
    text, n_pages, sha, nbytes, missing = None, 0, None, None, list(VERIFY_ANCHORS)
    if pdf is not None:
        sha = _sha256(pdf)
        nbytes = pdf.stat().st_size
        text, n_pages = extract_fulltext(pdf)
        if text:
            (proc / "action_plan_fulltext.txt").write_text(text, encoding="utf-8")
        missing = verify(text)
        if missing == VERIFY_ANCHORS and text is None:
            print("[action_plan] no PDF text reader available (pip install pdfplumber) - "
                  "skipped full text and verification; structured outputs still written")
        elif missing:
            print(f"[action_plan] WARNING: {len(missing)} verification anchor(s) not found "
                  f"in the PDF text: {missing}")
        else:
            print(f"[action_plan] verified: all {len(VERIFY_ANCHORS)} anchors present in the PDF text")
    else:
        print("[action_plan] no PDF at data/raw/action_plan/*.pdf - writing curated "
              "structured outputs without provenance / verification")

    wb_path = build_workbook(proc / WB_NAME)
    csvs = write_csvs(ext)

    manifest = {
        "source_key": "action_plan",
        "title": "Kenya Food Systems and Land Use Action Plan (2024-2030)",
        "publisher": "Government of Kenya, Ministry of Agriculture and Livestock Development; AGRA; FOLU Kenya",
        "date": "October 2023",
        "layer": "policy",
        "access": "open_download (manual)",
        "pdf_path": str(pdf) if pdf else None,
        "sha256": sha,
        "bytes": nbytes,
        "pages": n_pages,
        "verification_anchors": len(VERIFY_ANCHORS),
        "anchors_missing": missing,
        "workbook": str(wb_path),
        "csv_outputs": [str(p) for p in csvs],
        "extracted_by": "Aboubacar HEMA",
        "extracted_at": datetime.now(timezone.utc).isoformat(),
    }
    (proc / "action_plan_provenance.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    if prov is not None and sha is not None:
        try:
            prov.record(
                layer="policy", source_key="action_plan",
                meta={"title": manifest["title"], "publisher": manifest["publisher"],
                      "access": "open_download", "license": "Government of Kenya",
                      "url": ""},
                local_path=str(pdf), sha256=sha, nbytes=nbytes,
                status="ok", message=f"{n_pages} pages; {len(csvs)} CSV tables written")
        except Exception as exc:  # noqa: BLE001
            print(f"[action_plan] provenance record skipped: {exc}")

    print(f"[action_plan] workbook -> {wb_path}")
    print(f"[action_plan] {len(csvs)} CSV tables -> {ext}")
    return manifest


if __name__ == "__main__":
    here = Path(__file__).resolve().parents[1]
    run(here)
